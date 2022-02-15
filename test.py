from openpyxl import load_workbook
wb = load_workbook('test.xlsx')
ws = wb.active
a = ws.cell(1 , 1, 20)
print(a.value)
wb.save('test.xlsx')